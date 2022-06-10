import os
from pathlib import Path
from typing import Union
import matplotlib.pyplot as plt
import math

from openpyxl import Workbook, load_workbook
import openpyxl.drawing
from openpyxl.chart.axis import ChartLines
from openpyxl.chart import (
    ScatterChart,
    Reference,
    Series,
)
from openpyxl.utils.units import points_to_pixels, pixels_to_EMU
from openpyxl.drawing.line import LineProperties, LineEndProperties
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.chart.text import RichText
from openpyxl.drawing.text import Paragraph, ParagraphProperties, CharacterProperties, Font

import pandas as pd

stepinfo_template = {
    'overshoot': None,
    'rise_time': None,
    'settling_time': None
}

model_separator = '__'


def calc_exp_k(rk:float, xk:float) -> float:
    return -math.log(rk)/xk

def calc_err(x1, x2) -> float:
    err = x1 - x2
    if x2 != 0:
        err /= x2
    elif x1 != 0:
        err /= x1
    else:
        err = 0
    return abs(err)


def calc_stepinfo(ys:list, y_base:float, error_band=0.05, ts:list=None):
    overshoot = ((max(ys) if y_base > 0 else min(ys))-y_base)/y_base*100 if y_base != 0 else None #max(ys)
    try:
        tr = ts[next(i for i in range(0,len(ys)-1) if (ys[i]-ys[0])/(y_base-ys[0])>=(1-error_band))]-ts[0] if ts else None
    except StopIteration:
        tr = None
    try:
        tp = ts[next(len(ys)-i for i in range(1,len(ys)+1) if ((ys[len(ys)-i]-ys[0])/(y_base-ys[0])<=1-error_band or (ys[len(ys)-i]-ys[0])/(y_base-ys[0])>=1+error_band))]-ts[0] if ts else None
    except StopIteration:
        tp = None
    info = dict(stepinfo_template)
    info['overshoot'] = overshoot
    info['settling_time'] = tp
    info['rise_time'] = tr
    info['static_error'] = abs(ys[-1]-y_base) #)/y_base*100 if y_base != 0 else abs(ys[-1]-y_base)
    return info


class Integrator:
    def __init__(self, y0:float=0):
        self.y0 = y0
        self.sum = self.y0

    def reset(self):
        self.sum = self.y0

    def input(self, y:float):
        self.sum += y

    def output(self):
        return self.sum


class Derivative:
    def __init__(self, y0:float, dt_static=None):
        self.y0 = None
        self.y1 = y0
        self.var_step = dt_static is None
        self.dt = dt_static

    def input(self, y:float, dt=None):
        if self.var_step:
            if dt is None:
                raise ValueError("Нет информации о переменном шаге.")
            self.dt = dt
        self.y1, self.y0 = y, self.y1
        
    def output(self):
        if self.y0 is None:
            return self.y1/self.dt
        else:
            return (self.y1-self.y0)/self.dt


class DerivativeDict:
    def __init__(self):
        self.derivatives = {}

    def output(self, key:str):
        if key in self.derivatives:
            return self.derivatives[key].output()
        else:
            return 0
            #raise ValueError("Значения производной данного ключа отсутствует.")

    def input(self, key:str, y:float, dt:float):
        if key in self.derivatives:
            return self.derivatives[key].input(y)
        else:
            self.derivatives[key] = Derivative(y, dt_static=dt)
            return self.derivatives[key].output()


class MemoryDict:
    def __init__(self):
        self.memory = {}

    def input(self, key:str, y:float):
        self.memory[key] = y

    def output(self, key:str):
        return self.memory[key]


label_units = {
    'h': 'м',
    'U': 'В',
    'vartheta': 'град',
    'alpha': 'град',
    'wz': '1/с',
    'rew': '-',
    'deltaz': 'град',
    'x': 'м',
    'y': 'м',
    'V': 'м/с',
    'ax': 'м/с^2',
    'ay': 'м/с^2',
    't': 'с',
}

lineColors = [
    'ff0000', # красный
    'ff00d5', # розовый
    '4000ff', # синий
    '00fbff', # голубой
    '7f18c4', # фиолетовый
    '00ff37', # светло-зеленый
    'e6ff00', # желтый
    'ff7b00', # оранжевый
    '707070', # серый 
    'e875ff', # лиловый
    'ffc875', # типа персиковый
    'ffa6a6', # светло-розовый
    '000000', # черный \/
    '4a3d29', # коричневый \/
    '695c0e', # цвет детской неожиданности \/
    '100a38', # темно-синий \/
    '7d1d02', # бардовый \/
    '064720', # темно-зеленый \/
]

lineDashStyles = [
    None,
    'sysDot'
]

def comp_begin(label:str, target:str) -> bool:
    return len(label) >= len(target) and label[:len(target)] == target


def get_label_unit(label:str) -> Union[str, None]:
    for target, unit in label_units.items():
        if comp_begin(label, target):
            return f"[{unit}]"
    return None


def get_model_name_desc(model_name:str) -> str:
    description = ""
    method_to_name_mapping = {
        'obs': {
            "SPEED_MODE": "ПСР",
            "PID_SPEED_AERO": "ПСРА",
            'PID_LIKE': "Подобие",
        },
        'ctrl_mode':
        {
            "ADD_DIRECT_CONTROL": "ПКД",
            "ADD_PROC_CONTROL": "ОКД",
            "DIRECT_CONTROL": "ПУ",
        },
        'reset_ref_modes': {
            "CONST": "ПТУ",
            "OSCILLATING": "ОЗУ",
            "HYBRID": "ГИ",
        },
        'disturbance':
        {
            'AERO_DISTURBANCE': "Погрешность а/д"
        }
    }
    for mapping in method_to_name_mapping.values():
        for name, desc in mapping.items():
            if name in model_name:
                description += ' + ' + desc if description else desc
                model_name = model_name.replace(name, "")
                break
    if not description:
        description = model_name.split(model_separator)[-1]
    return description


def get_label_desc(label:str, index:int=None) -> str:
    if comp_begin(label, 'vartheta_ref'):
        return "Требуемый угол тангажа"
    elif comp_begin(label, "hzh"):
        return "Требуемая высота полета"
    elif model_separator not in label:
        return "СС ПИД"
    elif index is not None:
        return f'Конфигурация {index}'
    return get_model_name_desc(label)


def write_dataframe(data:pd.DataFrame, filename:str, bigMode=False):
    writer = pd.ExcelWriter(filename)
    data.to_excel(writer, index=True, header=True, sheet_name="data")

    ws = writer.sheets['data']

    def gather_styles(labels:list) -> dict:
        styles = []
        for i in range(len(lineColors)):
            for j in range(len(lineDashStyles)):
                styles.append({'color': lineColors[i], 'dashStyle': lineDashStyles[j]})
        info = {}
        for i in range(len(labels)):
            label = labels[i]
            info[label] = styles[i]
        return info

    def write_chart(labels:list, pos:str, wide_mode=True):
        styles = gather_styles(labels)
        chart = ScatterChart()
        fontName, fontSize = 'Times New Roman', (4000 if bigMode else 1400)
        font = Font(typeface=fontName)
        size = fontSize # 20 point size
        cp = CharacterProperties(latin=font, sz=size, b=False) # Try bold text
        pp = ParagraphProperties(defRPr=cp)
        rtp = RichText(p=[Paragraph(pPr=pp, endParaRPr=cp)])
        def formAxis(axis):
            width=points_to_pixels(7 if bigMode else 2)
            width=pixels_to_EMU(width)
            lineProp = LineProperties(w=width, solidFill = '000000', tailEnd=LineEndProperties(type='arrow', len='med'))
            prop = GraphicalProperties(ln=lineProp)
            axis.spPr = prop
            return axis
        chart.x_axis.title = data.index.name
        chart.x_axis.minorGridlines = ChartLines()
        chart.x_axis.txPr = rtp
        name = labels[0]
        chart.y_axis.title = name.split(model_separator)[0] if model_separator in name else name
        chart.y_axis.minorGridlines = ChartLines()
        chart.y_axis.txPr = rtp
        chart.x_axis, chart.y_axis = formAxis(chart.x_axis), formAxis(chart.y_axis)
        chart.x_axis.title.tx.rich.p[0].r[0].rPr = chart.y_axis.title.tx.rich.p[0].r[0].rPr = cp
        chart.legend.position = 'b'
        chart.legend.textProperties = rtp
        time = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row)
        for i in range(len(labels)):
            label = labels[i]
            style = styles[label]
            values = Reference(ws, min_col=data.columns.get_loc(label)+2, min_row=2, max_row=ws.max_row)
            series = Series(values, time, title_from_data=False, title=get_label_desc(label, index=i+1))
            if wide_mode:
                width=points_to_pixels(7 if bigMode else 2)
                width=pixels_to_EMU(width)
                lineProp = LineProperties(solidFill = style['color'], w=width)
            else:
                lineProp = LineProperties(solidFill = style['color'])
            series.graphicalProperties.line = lineProp
            dashStyle = style['dashStyle']
            if dashStyle:
                series.graphicalProperties.line.dashStyle = dashStyle
            series.smooth = True
            chart.series.append(series)
        ws.add_chart(chart, pos)

    groups = {}
    for column in data.columns:
        if model_separator in column:
            name = column.split(model_separator)[0]
        else:
            name = column
        if name in groups:
            groups[name].append(column)
        else:
            groups[name] = [column]
    
    for labels in groups.values():
        if labels[0] == 'vartheta, [град]':
            labels.append('vartheta_ref, [град]')
        elif labels[0] in ['h, [м]', 'y, [м]']:
            labels.append('hzh, [м]')
        write_chart(labels, 'A5')

    writer.save()


class Storage:
    def __init__(self):
        self.storage = {}
    
    def record(self, name, value):
        if name not in self.storage:
            self.storage[name] = []
        self.storage[name].append(value)
    
    def clear(self, name):
        del self.storage[name]

    def clear_all(self, ):
        self.storage = {}

    def plot(self, names, base:str=None, xlabel=None, ylabel=None):
        if type(names) is str:
            names = [names]
        for name in names:
            if base and base in self.storage:
                plt.plot(self.storage[base], self.storage[name], label=name)
            else:
                plt.plot(self.storage[name], label=name)
        plt.grid()
        plt.legend()
        if xlabel:
            plt.xlabel(xlabel)
        if ylabel:
            plt.ylabel(ylabel)
        plt.show()

    def save(self, filename='storage.xlsx', base=None):
        print(f"Сохраняю хранилище в {filename}")
        if len(self.storage) == 0:
            raise ValueError("Невозможно сохранить хранилище: пустое хранилище")
        os.makedirs(os.path.dirname(filename), exist_ok=True)
        data = pd.DataFrame.from_dict(self.storage, orient='columns')

        def place_unit(label:str) -> str:
            if model_separator in label:
                parts = label.split(model_separator)
                unit = get_label_unit(parts[0])
                if unit:
                    parts[0] = f"{parts[0]}, {unit}"
                return model_separator.join(parts)
            else:
                unit = get_label_unit(label)
                if unit:
                    return f"{label}, {unit}"
                else:
                    return label

        data.columns = list(map(place_unit, data.columns))
        if base and base in self.storage:
            data.set_index(place_unit(base), inplace=True)
        write_dataframe(data, filename)
        path = Path(filename)
        bigPath = os.path.dirname(path)/(path.stem+'_big'+path.suffix)
        write_dataframe(data, bigPath, bigMode=True)

    def set_suffix(self, suffix:str):
        self.storage = {f'{k}{model_separator}{suffix}': v for k, v in self.storage.items()}

    def merge(self, obj:'Storage', suffix:str):
        self.storage.update({f"{k}{model_separator}{suffix}": v for k, v in obj.storage.items()})