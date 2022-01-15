import matplotlib.pyplot as plt
from openpyxl import Workbook

stepinfo_template = {
    'overshoot': None,
    'rise_time': None,
    'settling_time': None
}

def calc_err(x1, x2) -> float:
    err = x1 - x2
    if x2 != 0:
        err /= x2
    elif x1 != 0:
        err /= x1
    else:
        err = 0
    return abs(err)

def calc_stepinfo(ys:list, y_base:float, error_band=0.05, ts:list=None):
    overshoot = (max(ys)-y_base)/y_base*100 if y_base != 0 else None #max(ys)
    try:
        tr = ts[next(i for i in range(0,len(ys)-1) if (ys[i]-ys[0])/(y_base-ys[0])>=(1-error_band))]-ts[0] if ts else None
    except StopIteration:
        tr = None
    try:
        tp = ts[next(len(ys)-i for i in range(1,len(ys)+1) if ((ys[len(ys)-i]-ys[0])/(y_base-ys[0])<=1-error_band or (ys[len(ys)-i]-ys[0])/(y_base-ys[0])>=1+error_band))]-ts[0] if ts else None
    except StopIteration:
        tp = None
    info = dict(stepinfo_template)
    info['overshoot'] = overshoot
    info['settling_time'] = tp
    info['rise_time'] = tr
    return info


class Integrator:
    def __init__(self, y0:float=0):
        self.y0 = y0
        self.sum = self.y0

    def reset(self):
        self.sum = self.y0

    def input(self, y:float):
        self.sum += y

    def output(self):
        return self.sum


class Derivative:
    def __init__(self, y0:float, dt_static=None):
        self.y0 = None
        self.y1 = y0
        self.var_step = dt_static is None
        self.dt = dt_static

    def input(self, y:float, dt=None):
        if self.var_step:
            if dt is None:
                raise ValueError("Нет информации о переменном шаге.")
            self.dt = dt
        self.y1, self.y0 = y, self.y1
        
    def output(self):
        if self.y0 is None:
            return self.y1/self.dt
        else:
            return (self.y1-self.y0)/self.dt


class DerivativeDict:
    def __init__(self):
        self.derivatives = {}

    def output(self, key:str):
        if key in self.derivatives:
            return self.derivatives[key].output()
        else:
            return 0
            #raise ValueError("Значения производной данного ключа отсутствует.")

    def input(self, key:str, y:float, dt:float):
        if key in self.derivatives:
            return self.derivatives[key].input(y)
        else:
            self.derivatives[key] = Derivative(y, dt_static=dt)
            return self.derivatives[key].output()


class MemoryDict:
    def __init__(self):
        self.memory = {}

    def input(self, key:str, y:float):
        self.memory[key] = y

    def output(self, key:str):
        return self.memory[key]


class Storage:
    def __init__(self):
        self.storage = {}
    
    def record(self, name, value):
        if name not in self.storage:
            self.storage[name] = []
        self.storage[name].append(value)
    
    def clear(self, name):
        del self.storage[name]

    def clear_all(self, ):
        self.storage = {}

    def plot(self, names, base:str=None):
        if type(names) is str:
            names = [names]
        for name in names:
            if base and base in self.storage:
                plt.plot(self.storage[base], self.storage[name], label=name)
            else:
                plt.plot(self.storage[name], label=name)
        plt.grid()
        plt.legend()
        plt.show()

    def save(self, filename='storage.xls', base=None):
        if len(self.storage) == 0:
            raise ValueError("Невозможно сохранить хранилище: пустое хранилище")
        wb = Workbook()
        ws = wb.active
        if base:
            j = 2
            ws.cell(row=1, column=1).value = base
            for i in range(len(self.storage[base])):
                ws.cell(row=i+2, column=1).value = self.storage[base][i]
        else:
            j = 1
        for k, v in self.storage.items():
            if base and k == base:
                continue
            ws.cell(row=1, column=j).value = k
            for i in range(len(v)):
                ws.cell(row=i+2, column=j).value = v[i]
            j += 1
        wb.save(filename)

    def merge(self, obj, prefix:str):
        self.storage.update(dict([(k+'_'+prefix, v) for k,v in obj.storage.items()]))